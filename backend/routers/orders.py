from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, date
from io import BytesIO, StringIO
import csv
import database
from routers.auth import verify_token
import httpx
import os
import config

router = APIRouter()

# Pydantic models
class OrderBase(BaseModel):
    id: int
    user_id: int
    username: Optional[str]
    full_name: str
    status: str
    order_data: Optional[dict] = None
    photo_file_id: Optional[str]
    created_at: datetime
    internal_note: Optional[str]
    price: Optional[float] = None
    price_currency: Optional[str] = None
    price_status: Optional[str] = None
    deadline: Optional[date] = None

class OrderUpdate(BaseModel):
    status: Optional[str] = None
    internal_note: Optional[str] = None
    price: Optional[float] = None
    currency: Optional[str] = None
    deadline: Optional[date] = None

class OrderStats(BaseModel):
    total_orders: int
    new_orders: int
    active_orders: int

class PriceSet(BaseModel):
    price: float
    currency: str = "UAH"

DEFAULT_PAGE_LIMIT = 20


@router.get("/", response_model=List[OrderBase])
async def get_orders(
    page: int = 1,
    limit: int = DEFAULT_PAGE_LIMIT,
    status_filter: Optional[str] = None,
    _payload: dict = Depends(verify_token)
):
    """Получить список заказов с пагинацией"""
    try:
        offset = (page - 1) * limit
        orders = database.get_orders_paginated(limit, offset, status_filter)
        return orders
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching orders: {str(e)}")

@router.get("/stats", response_model=OrderStats)
async def get_order_stats(_payload: dict = Depends(verify_token)):
    """Получить статистику заказов"""
    try:
        stats = database.get_order_statistics()
        return stats
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching statistics: {str(e)}")


@router.get("/export/csv")
async def export_orders_csv(
    status_filter: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    _payload: dict = Depends(verify_token)
):
    """Export orders as CSV file"""
    try:
        orders = database.get_orders_for_export(status_filter, date_from, date_to)
        output = StringIO()
        writer = csv.writer(output)

        # Collect all unique order_data keys across all orders for extra columns
        order_data_keys = set()
        for order in orders:
            od = order.get('order_data') or {}
            order_data_keys.update(od.keys())
        order_data_keys = sorted(order_data_keys)

        base_headers = [
            "ID", "Client", "Username", "Status", "Price", "Currency",
            "Deadline", "Created At"
        ]
        headers = base_headers + order_data_keys
        writer.writerow(headers)

        for order in orders:
            od = order.get('order_data') or {}
            row = [
                order.get('id'),
                order.get('full_name', ''),
                order.get('username', ''),
                order.get('status', ''),
                order.get('price', ''),
                order.get('price_currency', ''),
                str(order.get('deadline', '')) if order.get('deadline') else '',
                str(order.get('created_at', '')),
            ]
            for key in order_data_keys:
                row.append(od.get(key, ''))
            writer.writerow(row)

        output.seek(0)
        filename = f"orders_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting CSV: {str(e)}")


@router.get("/export/excel")
async def export_orders_excel(
    status_filter: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    _payload: dict = Depends(verify_token)
):
    """Export orders as Excel file"""
    try:
        from openpyxl import Workbook

        orders = database.get_orders_for_export(status_filter, date_from, date_to)

        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"

        # Collect all unique order_data keys
        order_data_keys = set()
        for order in orders:
            od = order.get('order_data') or {}
            order_data_keys.update(od.keys())
        order_data_keys = sorted(order_data_keys)

        base_headers = [
            "ID", "Client", "Username", "Status", "Price", "Currency",
            "Deadline", "Created At"
        ]
        headers = base_headers + order_data_keys
        ws.append(headers)

        for order in orders:
            od = order.get('order_data') or {}
            row = [
                order.get('id'),
                order.get('full_name', ''),
                order.get('username', ''),
                order.get('status', ''),
                order.get('price', ''),
                order.get('price_currency', ''),
                str(order.get('deadline', '')) if order.get('deadline') else '',
                str(order.get('created_at', '')),
            ]
            for key in order_data_keys:
                row.append(od.get(key, ''))
            ws.append(row)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"orders_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting Excel: {str(e)}")


@router.get("/analytics")
async def get_analytics(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    _payload: dict = Depends(verify_token)
):
    """Get analytics data for orders"""
    try:
        analytics = database.get_analytics(date_from, date_to)
        return analytics
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching analytics: {str(e)}")


@router.get("/{order_id}", response_model=OrderBase)
async def get_order(order_id: int, _payload: dict = Depends(verify_token)):
    """Получить заказ по ID"""
    try:
        order = database.get_order(order_id)
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        return order
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching order: {str(e)}")


@router.get("/{order_id}/invoice")
async def get_order_invoice(
    order_id: int,
    _payload: dict = Depends(verify_token)
):
    """Generate PDF invoice for an order"""
    try:
        from fpdf import FPDF

        order = database.get_order(order_id)
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 20)
        pdf.cell(0, 15, f"Invoice #{order_id}", ln=True, align="C")

        pdf.set_font("Helvetica", "", 11)
        pdf.cell(0, 8, f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}", ln=True)

        pdf.ln(5)
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 10, "Customer Information", ln=True)
        pdf.set_font("Helvetica", "", 11)
        pdf.cell(0, 7, f"Name: {order.get('full_name', 'N/A')}", ln=True)
        username = order.get('username')
        if username:
            pdf.cell(0, 7, f"Username: @{username}", ln=True)
        pdf.cell(0, 7, f"User ID: {order.get('user_id', 'N/A')}", ln=True)

        pdf.ln(5)
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 10, "Order Details", ln=True)
        pdf.set_font("Helvetica", "", 11)
        pdf.cell(0, 7, f"Order ID: {order_id}", ln=True)
        pdf.cell(0, 7, f"Status: {order.get('status', 'N/A')}", ln=True)

        order_data = order.get('order_data') or {}
        if order_data:
            pdf.ln(3)
            for key, value in order_data.items():
                label = key.replace('_', ' ').title()
                pdf.cell(0, 7, f"{label}: {value}", ln=True)

        pdf.ln(5)
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 10, "Pricing", ln=True)
        pdf.set_font("Helvetica", "", 11)
        price = order.get('price')
        currency = order.get('price_currency', 'UAH')
        if price:
            pdf.cell(0, 7, f"Price: {price} {currency}", ln=True)
        else:
            pdf.cell(0, 7, "Price: Not set", ln=True)

        deadline = order.get('deadline')
        if deadline:
            pdf.ln(3)
            pdf.cell(0, 7, f"Deadline: {deadline}", ln=True)

        buffer = BytesIO()
        pdf.output(buffer)
        buffer.seek(0)

        filename = f"invoice_{order_id}.pdf"
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating invoice: {str(e)}")


@router.put("/{order_id}")
async def update_order(
    order_id: int,
    order_update: OrderUpdate,
    _payload: dict = Depends(verify_token)
):
    """Обновить заказ"""
    try:
        current_order = database.get_order(order_id)
        if not current_order:
            raise HTTPException(status_code=404, detail="Order not found")

        old_status = current_order.get('status')

        if order_update.status:
            # When status changes to 'done', use mark_order_completed which sets completed_at
            if old_status != order_update.status and order_update.status == 'done':
                database.mark_order_completed(order_id)
            else:
                database.update_order_field(order_id, 'status', order_update.status)

            # Если статус изменился - отправляем уведомление
            if old_status != order_update.status:
                await send_status_update_notification(
                    current_order['user_id'],
                    order_id,
                    order_update.status
                )

        if order_update.internal_note is not None:
            database.update_order_field(order_id, 'internal_note', order_update.internal_note)

        if order_update.price is not None:
            old_price = current_order.get('price')
            new_price = order_update.price
            new_currency = order_update.currency or 'UAH'
            
            if old_price != new_price or current_order.get('price_currency') != new_currency:
                database.set_order_price(order_id, new_price, new_currency)
                # Уведомляем клиента о новой цене
                await send_price_notification(
                    current_order['user_id'],
                    order_id,
                    new_price,
                    new_currency
                )

        if order_update.deadline is not None:
            if current_order.get('deadline') != order_update.deadline:
                database.set_order_deadline(order_id, order_update.deadline)

        return {"message": "Order updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Update order error: {e}")
        raise HTTPException(status_code=500, detail=f"Error updating order: {str(e)}")


@router.put("/{order_id}/price")
async def set_order_price(
    order_id: int,
    price_data: PriceSet,
    _payload: dict = Depends(verify_token)
):
    """Set price on an order and notify client via Telegram"""
    try:
        order = database.get_order(order_id)
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")

        database.set_order_price(order_id, price_data.price, price_data.currency)

        # Send price notification to client with inline buttons
        await send_price_notification(
            order['user_id'],
            order_id,
            price_data.price,
            price_data.currency
        )

        return {"message": "Price set successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error setting price: {str(e)}")


@router.delete("/all")
async def delete_all_orders(_payload: dict = Depends(verify_token)):
    """Удалить все заказы"""
    try:
        database.delete_all_orders()
        return {"message": "All orders deleted"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting orders: {str(e)}")


async def send_status_update_notification(user_id: int, order_id: int, new_status: str):
    """Send status update notification to client via Telegram Bot API"""
    lang = database.get_user_language(user_id)

    status_labels = {
        'new': {'ru': 'НОВЫЙ (Принят в обработку)', 'en': 'NEW (Accepted)', 'uk': 'НОВЕ (Прийнято)'},
        'discussion': {'ru': 'Обсуждение деталей', 'en': 'Discussion', 'uk': "Обговорення деталей"},
        'approved': {'ru': 'Одобрен / В работе', 'en': 'Approved / In Progress', 'uk': 'Схвалено / В роботі'},
        'work': {'ru': 'Выполняется', 'en': 'In Progress', 'uk': 'Виконується'},
        'done': {'ru': 'ГОТОВ!', 'en': 'DONE!', 'uk': 'ГОТОВО!'},
        'rejected': {'ru': 'Отказ', 'en': 'Rejected', 'uk': 'Відмова'},
    }

    status_text = status_labels.get(new_status, {}).get(lang, new_status)

    headers_map = {
        'ru': f'<b>Статус вашего заказа #{order_id} изменен:</b>',
        'en': f'<b>Order #{order_id} status changed:</b>',
        'uk': f'<b>Статус вашого замовлення #{order_id} змінено:</b>',
    }
    message_text = f"{headers_map.get(lang, headers_map['ru'])}\n\n{status_text}"

    bot_token = config.BOT_TOKEN
    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"

    async with httpx.AsyncClient() as client:
        try:
            payload = {
                "chat_id": user_id,
                "text": message_text,
                "parse_mode": "HTML"
            }
            response = await client.post(url, json=payload)
            response.raise_for_status()
        except Exception as e:
            print(f"Failed to send Telegram notification: {e}")


async def send_price_notification(user_id: int, order_id: int, price: float, currency: str):
    """Send price notification to client with inline accept/decline buttons"""
    bot_token = config.BOT_TOKEN
    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    lang = database.get_user_language(user_id)

    messages = {
        'ru': f"<b>Установлена цена по заказу #{order_id}</b>\n\nСумма: <b>{price} {currency}</b>\n\nПожалуйста, подтвердите или отклоните:",
        'en': f"<b>Price set for order #{order_id}</b>\n\nAmount: <b>{price} {currency}</b>\n\nPlease accept or decline:",
        'uk': f"<b>Встановлено ціну за замовлення #{order_id}</b>\n\nСума: <b>{price} {currency}</b>\n\nБудь ласка, підтвердіть або відхиліть:",
    }
    message_text = messages.get(lang, messages['ru'])

    btn_accept = {'ru': 'Принять', 'en': 'Accept', 'uk': 'Прийняти'}
    btn_reject = {'ru': 'Отклонить', 'en': 'Decline', 'uk': 'Відхилити'}

    inline_keyboard = {
        "inline_keyboard": [
            [
                {"text": btn_accept.get(lang, 'Принять'), "callback_data": f"price_accept:{order_id}"},
                {"text": btn_reject.get(lang, 'Отклонить'), "callback_data": f"price_reject:{order_id}"}
            ]
        ]
    }

    async with httpx.AsyncClient() as client:
        try:
            payload = {
                "chat_id": user_id,
                "text": message_text,
                "parse_mode": "HTML",
                "reply_markup": inline_keyboard
            }
            response = await client.post(url, json=payload)
            response.raise_for_status()
        except Exception as e:
            print(f"Failed to send price notification: {e}")


@router.get("/{order_id}/photos")
async def get_order_photos(order_id: int, _payload: dict = Depends(verify_token)):
    """Получить рабочие ссылки на фото заказа"""
    try:
        order = database.get_order(order_id)
        if not order or not order['photo_file_id']:
            return {"photos": []}

        raw_ids = order['photo_file_id'].split(',')
        # Очищаем ID от префиксов p: и d:
        clean_ids = [p[2:] if p.startswith(('p:', 'd:')) else p for p in raw_ids]

        photo_urls = []
        bot_token = config.BOT_TOKEN

        async with httpx.AsyncClient() as client:
            for file_id in clean_ids:
                try:
                    # 1. Получаем путь к файлу через getFile
                    file_info_url = f"https://api.telegram.org/bot{bot_token}/getFile?file_id={file_id}"
                    resp = await client.get(file_info_url)

                    if resp.status_code == 200:
                        file_path = resp.json().get('result', {}).get('file_path')
                        if file_path:
                            # 2. Формируем прямую ссылку на скачивание
                            full_url = f"https://api.telegram.org/file/bot{bot_token}/{file_path}"
                            photo_urls.append(full_url)
                except Exception as e:
                    print(f"Error resolving file_id {file_id}: {e}")

        return {"photos": photo_urls}
    except Exception as e:
        print(f"Get photos error: {e}")
        raise HTTPException(status_code=500, detail=f"Error fetching photos: {str(e)}")
